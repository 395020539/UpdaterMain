import xlrd
import openpyxl
from logging_maker import logger
# from get_file_path import *

# file_mech_table = "MechanicalDataSheet -机械参数表 Voyah_H97C_20220902.xlsm"
# from get_file_path import get_file_path
# file_a2l, file_geskon, file_dcm, file_data_mytable, file_mech_table = get_file_path()

# file_a2l = get_file_path.file_a2l
# file_dcm = get_file_path.file_dcm
# file_geskon = get_file_path.file_geskon
# file_mech_table = get_file_path.file_mech_table
# file_data_mytable = get_file_path.file_data_mytable

# workbook = xlrd.open_workbook(file_mech_table)

def check_mech_sheet(sheet_names):
    check_mech_sheet_result = True
    # if len(sheet_names) != 9:
    #     check_mech_sheet_result = False
    #     print("MechanicalDataSheet is not complete")
    #     logger.critical("MechanicalDataSheet is not complete")
    if not "MechanicalData" in sheet_names[1] or \
            not "Steering Ratio" in sheet_names[2] or \
            not "SW-Endstop" in sheet_names[3] or \
            not "Steering Angle" in sheet_names[4] or \
            not "SA_Config" in sheet_names[5] or \
            not "RackWheelConverter" in sheet_names[6] or \
            not "LoadTracking" in sheet_names[7] or \
            not "other tuning" in sheet_names[8]:
        print("MechanicalDataSheet is not complete")
        logger.critical("MechanicalDataSheet is not complete")
    print(f"check_mech_sheet_result = {check_mech_sheet_result}")
    logger.info(f"check_mech_sheet_result = {check_mech_sheet_result}")
    return check_mech_sheet_result

# 通过定位查找机械参数表
def fun_find_mech(file_mech_table, located_sheet, position_row, position_col, number_float):
    print(f"查找机械参数表:{located_sheet}，坐标: {position_row},{position_col}")
    logger.info(f"查找机械参数表:{located_sheet}，坐标: {position_row},{position_col}")
    # located_sheet 参数值所在表名
    # position_row 和 position_col 是参数值所在实际位置减 1
    # number_float 小数点后几位
    mech_value = ""
    find_result = False
    try:
        mech_work_book = xlrd.open_workbook(file_mech_table)
        sheet_names = mech_work_book.sheet_names()
        print(sheet_names)
        check_mech_sheet_result = check_mech_sheet(sheet_names)
        if check_mech_sheet_result:
            mech_work_sheet_tofind = mech_work_book.sheet_by_name(located_sheet)
            value_type = mech_work_sheet_tofind.cell_type(position_row, position_col)
            '''（0: 空，1: Unicode字符串，2: 数值，3: 日期，4: 布尔，5: 出错代码，6: 空）'''
            print(f"value type: {value_type}")
            logger.debug(f"value type: {value_type}")
            value_raw = mech_work_sheet_tofind.cell_value(position_row, position_col)
            if value_type == 1:
                mech_value = value_raw
                find_result = True
            elif value_type == 2:
                find_result = True
                if number_float is None:
                    mech_value = value_raw
                else:
                    mech_value = round(value_raw, number_float)
            else:
                mech_value = ""
            print(f"value_raw: {value_raw}, value: {mech_value}")
            logger.info(f"value_raw: {value_raw}, value: {mech_value}")
    except FileNotFoundError:
        print(f"No such file or directory")
        logger.critical(f"No such file or directory")
    except Exception as e:
        print(f"Other error occurred: {e}")
        logger.critical(f"Other error occurred: {e}")
    finally:
        print(f"机械参数表查询结果：{find_result}; 查询参数值: {mech_value}")
        logger.info(f"机械参数表查询结果：{find_result}; 查询参数值: {mech_value}")
        return find_result, mech_value


# find_result, mech_value = fun_find_mech(file_mech_table, "MechanicalData", 68, 1, 5)
# print(f"find_result: {find_result}; mech_value: {mech_value}")


# 通过parameter在指定sheet中查找机械参数值
def fun_find_mech_by_para(file_mech_table,search_parameter, search_in_sheet_list):
    global mech_workbook
    print(f"查找机械参数表参数:{search_parameter}")
    logger.info(f"查找机械参数表参数:{search_parameter}")
    try:
        # 打开 Excel 文件
        mech_workbook = openpyxl.load_workbook(file_mech_table, data_only=True)

        # 遍历所有工作表
        for sheetname in search_in_sheet_list:
            sheet = mech_workbook[sheetname]
            # 遍历所有单元格
            found_value = None
            for row in sheet.iter_rows():
                for cell in row:
                    # print(f'位置 {cell.coordinate} ')
                    # print(cell.value, type(cell.value))
                    # value = cell.value
                    # print("value:"+ str(value))
                    if cell and str(cell.value).strip() == search_parameter:
                        # 找到了目标值
                        print(f'Parameter "{search_parameter}" found at {cell.coordinate} in sheet {sheetname}')
                        logger.info(f'Parameter "{search_parameter}" found at {cell.coordinate} in sheet {sheetname}')

                        # 获取同一行中的另一个单元格的值
                        found_cell = sheet.cell(row=cell.row, column=cell.column - 2)
                        found_value = found_cell.value
                        # print(f'Other value in the same row is {found_value}')
                        if found_cell.data_type == 'f':
                            # 如果单元格包含公式，则获取其计算后的值
                            found_value = found_cell.value
                        else:
                            # 否则，获取常规文本或数字值
                            found_value = found_cell.value
                        print(f'Parameter value is {found_value}')
                        logger.info(f'Parameter value is {found_value}')
                        break
                if found_value is not None:
                    break
            if found_value is not None:
                break
        if found_value is None:
            found_value = ""
    except FileNotFoundError:
        print("File not found error: please check the file path and name.")
        logger.critical("File not found error: please check the file path and name.")
    except AttributeError:
        print("Attribute error: please check the sheet name.")
        logger.critical("Attribute error: please check the sheet name.")
    except IndexError:
        print("Index error: please check the sheet index.")
        logger.critical("Index error: please check the sheet index.")
    except Exception as e:
        print(f"Other error occurred: {e}")
        logger.critical(f"Other error occurred: {e}")
    finally:
        # 如果还没有关闭文件，关闭文件并释放资源
        if mech_workbook:
            mech_workbook.close()
        return found_value

# found_value = fun_find_mech_by_para(search_parameter, search_in_sheet_list)
# print(f"参数值为: {found_value}")
